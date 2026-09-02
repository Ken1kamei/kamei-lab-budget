from io import BytesIO
from pathlib import Path

from numbers_parser import Document
from openpyxl import Workbook, load_workbook

from labapps.services.gantt import (
    build_gantt_context,
    merge_project_gantt,
    parse_gantt_csv,
    parse_gantt_file,
    parse_gantt_numbers,
    parse_gantt_workbook,
    resolve_gantt_rows,
)


def vertex_style_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Project schedule"
    sheet.append(["", "TASK", "ASSIGNED TO", "PROGRESS", "START", "END"])
    sheet.append(["", "Planning", "", "", "", ""])
    sheet.append(["", "Define scope", "member@nyu.edu", 0.5, "09/01/2026", "09/05/2026"])
    sheet.append(["", "Run pilot", "Unknown Person", 100, "09/06/2026", "09/12/2026"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_vertex_style_gantt_workbook():
    result = parse_gantt_workbook(vertex_style_workbook())

    assert result.sheet_name == "Project schedule"
    assert result.errors == []
    assert len(result.rows) == 2
    assert result.rows[0] == {
        "source_row": 3,
        "milestone_id": "",
        "phase": "Planning",
        "task": "Define scope",
        "assigned_to": "member@nyu.edu",
        "progress_percent": 50.0,
        "start_date": "2026-09-01",
        "due_date": "2026-09-05",
        "status": "In progress",
        "next_action": "",
    }
    assert result.rows[1]["status"] == "Completed"


def test_parse_rejects_non_excel_content():
    result = parse_gantt_workbook(b"not an Excel workbook")

    assert result.rows == []
    assert result.errors
    assert "could not be read safely" in result.errors[0]


def test_parse_csv_exported_from_numbers_with_utf8_bom():
    content = (
        "\ufeffPhase;Task;Assigned to;Start Date;End Date;Progress %;Status;Next Action\n"
        "Planning;Define scope;member@nyu.edu;2026/09/01;2026/09/05;50%;In progress;Review scope\n"
        "Execution;Run pilot;;2026/09/06;2026/09/12;0;Not started;Run pilot\n"
    ).encode("utf-8")

    result = parse_gantt_csv(content, filename="project-gantt.csv")

    assert result.errors == []
    assert result.sheet_name == "project-gantt.csv"
    assert result.header_row == 1
    assert [row["task"] for row in result.rows] == ["Define scope", "Run pilot"]
    assert result.rows[0]["assigned_to"] == "member@nyu.edu"
    assert result.rows[0]["progress_percent"] == 50.0
    assert result.rows[0]["start_date"] == "2026-09-01"


def test_parse_apple_numbers_file(tmp_path):
    path = tmp_path / "project-gantt.numbers"
    document = Document(
        sheet_name="Gantt Import",
        table_name="Gantt Import",
        num_header_rows=1,
        num_header_cols=0,
        num_rows=3,
        num_cols=8,
    )
    table = document.sheets[0].tables[0]
    table_rows = [
        [
            "Phase",
            "Task",
            "Assigned to",
            "Start Date",
            "End Date",
            "Progress %",
            "Status",
            "Next Action",
        ],
        [
            "Planning",
            "Define scope",
            "member@nyu.edu",
            "2026-09-01",
            "2026-09-05",
            50,
            "In progress",
            "Review scope",
        ],
        [
            "Execution",
            "Run pilot",
            "",
            "2026-09-06",
            "2026-09-12",
            0,
            "Not started",
            "Run pilot",
        ],
    ]
    for row_number, row in enumerate(table_rows):
        for column_number, value in enumerate(row):
            table.write(row_number, column_number, value)
    document.save(path)

    result = parse_gantt_numbers(path.read_bytes())

    assert result.errors == []
    assert result.sheet_name == "Gantt Import / Gantt Import"
    assert result.header_row == 1
    assert [row["task"] for row in result.rows] == ["Define scope", "Run pilot"]
    assert result.rows[0]["assigned_to"] == "member@nyu.edu"
    assert result.rows[0]["progress_percent"] == 50.0


def test_parse_gantt_file_dispatches_by_uploaded_filename(tmp_path):
    csv_path = tmp_path / "project-gantt.csv"
    csv_path.write_text(
        "Task,Start Date,End Date\nCSV task,2026-09-01,2026-09-03\n",
        encoding="utf-8",
    )

    with csv_path.open("rb") as uploaded:
        result = parse_gantt_file(uploaded)

    assert result.errors == []
    assert result.rows[0]["task"] == "CSV task"


def test_distributed_template_has_dynamic_calendar_and_remains_importable():
    template_path = (
        Path(__file__).resolve().parents[2]
        / "labapps"
        / "static"
        / "labapps"
        / "Kamei_Lab_Gantt_Import_Template.xlsx"
    )
    exact_result = parse_gantt_workbook(template_path.read_bytes())

    assert exact_result.header_row == 7
    assert exact_result.rows == []
    assert exact_result.errors == [
        "The Gantt table does not contain any importable task rows."
    ]

    workbook = load_workbook(template_path, data_only=False)
    sheet = workbook["Gantt Import"]

    assert sheet["D2"].value == '=IF(COUNT(NP8:NP37)=0,"",MIN(NP8:NP37))'
    assert sheet["F2"].value == '=IF(COUNT(NQ8:NQ37)=0,"",MAX(NQ8:NQ37))'
    assert sheet["D3"].value == '=IF(D2="","",D2-WEEKDAY(D2,2)+1)'
    assert sheet["F3"].value == '=IF(F2="","",F2)'
    assert sheet["I6"].value == '=IF($D$3="","",$D$3)'
    assert sheet["J6"].value == '=IF(OR(I6="",I6>=$F$2),"",I6+1)'
    assert (
        sheet["I7"].value
        == '=IF(I6="","",CHOOSE(WEEKDAY(I6,2),"M","T","W","T","F","S","S"))'
    )
    assert (
        sheet["NP8"].value
        == '=IF(D8="","",IF(ISNUMBER(D8),D8,IFERROR(DATE(LEFT(D8,4),MID(D8,6,2),RIGHT(D8,2)),"")))'
    )
    assert (
        sheet["NQ8"].value
        == '=IF(E8="","",IF(ISNUMBER(E8),E8,IFERROR(DATE(LEFT(E8,4),MID(E8,6,2),RIGHT(E8,2)),"")))'
    )
    assert sheet["NO6"].value == '=IF(OR(NN6="",NN6>=$F$2),"",NN6+1)'
    assert "I5:O5" in {str(cell_range) for cell_range in sheet.merged_cells.ranges}
    assert "NI5:NO5" in {
        str(cell_range) for cell_range in sheet.merged_cells.ranges
    }

    conditional_formatting = list(sheet.conditional_formatting)
    assert [str(item.sqref) for item in conditional_formatting] == ["I8:NO37"]
    rules = sheet.conditional_formatting[conditional_formatting[0]]
    assert [rule.priority for rule in rules] == [1, 2]
    assert "$NP8" in rules[0].formula[0]
    assert "$NQ8" in rules[0].formula[0]
    assert "$NP8" in rules[1].formula[0]
    assert "$NQ8" in rules[1].formula[0]

    sheet["A8"] = "Planning"
    sheet["B8"] = "Dynamic template import check"
    sheet["C8"] = "member@nyu.edu"
    sheet["D8"] = "2026-09-01"
    sheet["E8"] = "2026-09-05"
    sheet["F8"] = 50
    sheet["G8"] = "In progress"
    sheet["A29"] = "Reporting"
    sheet["B29"] = "Lower template row import check"
    sheet["C29"] = "member@nyu.edu"
    sheet["D29"] = "2026-11-02"
    sheet["E29"] = "2026-11-06"
    sheet["F29"] = 0
    sheet["G29"] = "Not started"
    buffer = BytesIO()
    workbook.save(buffer)

    result = parse_gantt_workbook(buffer.getvalue())

    assert result.errors == []
    assert result.header_row == 7
    assert len(result.rows) == 2
    assert result.rows[0]["task"] == "Dynamic template import check"
    assert result.rows[0]["start_date"] == "2026-09-01"
    assert result.rows[0]["due_date"] == "2026-09-05"
    assert result.rows[1]["source_row"] == 29
    assert result.rows[1]["task"] == "Lower template row import check"


def test_resolve_and_merge_gantt_rows_are_idempotent_and_preserve_manual_rows():
    parsed = parse_gantt_workbook(vertex_style_workbook())
    project = {"project_id": "P001", "project": "Chip study", "aim": "Model disease"}
    members = [
        {
            "member_id": "M001",
            "email": "member@nyu.edu",
            "name": "Lab Member",
            "display_name": "Member",
        },
        {"member_id": "M002", "email": "lead@nyu.edu", "display_name": "Lead"},
    ]

    first, warnings = resolve_gantt_rows(
        parsed.rows,
        project=project,
        members=members,
        default_owner_member_id="M002",
        updated_at="2026-07-23T12:00:00+04:00",
    )
    second, _ = resolve_gantt_rows(
        parsed.rows,
        project=project,
        members=members,
        default_owner_member_id="M002",
        updated_at="2026-07-23T12:00:00+04:00",
    )

    assert [row["milestone_id"] for row in first] == [
        row["milestone_id"] for row in second
    ]
    assert first[0]["owner_member_id"] == "M001"
    assert first[1]["owner_member_id"] == "M002"
    assert "Unknown Person" in warnings[0]

    existing = [
        {"milestone_id": "MS001", "project_id": "P001", "milestone": "Manual"},
        {"milestone_id": "MS-GANTT-OLD", "project_id": "P001", "milestone": "Old"},
        {"milestone_id": "MS-GANTT-OTHER", "project_id": "P002", "milestone": "Other"},
    ]
    merged = merge_project_gantt(existing, "P001", first)
    ids = {row["milestone_id"] for row in merged}

    assert "MS001" in ids
    assert "MS-GANTT-OLD" not in ids
    assert "MS-GANTT-OTHER" in ids
    assert {row["milestone_id"] for row in first}.issubset(ids)


def test_build_gantt_context_uses_project_specific_rows():
    project = {"project_id": "P001", "project": "Chip study"}
    milestones = [
        {
            "milestone_id": "MS1",
            "project_id": "P001",
            "milestone": "Pilot",
            "time_window": "Setup",
            "owner_member_id": "M001",
            "start_date": "2026-09-01",
            "due_date": "2026-09-10",
            "progress_percent": "60",
            "status": "In progress",
        },
        {
            "milestone_id": "MS2",
            "project_id": "P002",
            "milestone": "Outside scope",
            "start_date": "2026-08-01",
            "due_date": "2026-12-01",
        },
    ]

    gantt = build_gantt_context(project, milestones, {"M001": "Ken"})

    assert len(gantt["rows"]) == 1
    assert gantt["rows"][0]["milestone"] == "Pilot"
    assert gantt["rows"][0]["owner"] == "Ken"
    assert gantt["rows"][0]["progress_width"] == 60.0
    assert gantt["start_date"].isoformat() == "2026-09-01"
    assert gantt["end_date"].isoformat() == "2026-09-10"
