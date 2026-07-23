from __future__ import annotations

import hashlib
import io
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.datetime import from_excel


HEADER_ALIASES = {
    "milestone_id": {"milestoneid", "taskid", "id"},
    "phase": {"phase", "stage", "workstream", "category", "timewindow"},
    "task": {"task", "taskname", "milestone", "activity", "deliverable"},
    "assigned_to": {
        "assignedto",
        "assignee",
        "owner",
        "researcher",
        "lead",
        "personresponsible",
    },
    "progress": {"progress", "progresspercent", "completion", "percentcomplete"},
    "start": {"start", "startdate", "plannedstart"},
    "end": {"end", "enddate", "due", "duedate", "targetdate", "plannedend"},
    "status": {"status", "taskstatus"},
    "next_action": {"nextaction", "nextstep", "action"},
}
REQUIRED_FIELDS = {"task", "start", "end"}
STATUS_VALUES = {"Not started", "In progress", "Blocked", "Completed"}
IGNORED_TASK_PREFIXES = (
    "do not delete",
    "insert new rows",
    "example -",
    "example:",
)
MAX_WORKBOOK_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_ARCHIVE_MEMBERS = 2_000
MAX_WORKSHEETS = 20
MAX_EMPTY_TASK_ROWS = 100


@dataclass
class GanttImportResult:
    sheet_name: str
    header_row: int
    rows: list[dict]
    warnings: list[str]
    errors: list[str]


def _header_key(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).casefold()
    return re.sub(r"[^a-z0-9]+", "", text)


def _field_for_header(value) -> str:
    key = _header_key(value)
    return next(
        (field for field, aliases in HEADER_ALIASES.items() if key in aliases),
        "",
    )


def _parse_date(value, workbook) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, workbook.epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _parse_progress(value) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        text = value.strip().replace("%", "")
        try:
            number = float(text)
        except ValueError:
            return None
        return max(0.0, min(100.0, number))
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if 0 <= number <= 1:
        number *= 100
    return max(0.0, min(100.0, number))


def _status(value, progress: float | None) -> str:
    raw = str(value or "").strip()
    normalized = {item.casefold(): item for item in STATUS_VALUES}
    if raw.casefold() in normalized:
        return normalized[raw.casefold()]
    if progress is not None and progress >= 100:
        return "Completed"
    if progress is not None and progress > 0:
        return "In progress"
    return "Not started"


def _find_header(workbook):
    preferred = sorted(
        workbook.worksheets,
        key=lambda sheet: sheet.title != "Gantt Import",
    )
    for sheet in preferred:
        header_rows = sheet.iter_rows(
            min_row=1,
            max_row=40,
            max_col=min(sheet.max_column or 64, 256),
            values_only=True,
        )
        for row_number, header_values in enumerate(header_rows, start=1):
            mapping = {}
            for column_number, cell in enumerate(
                header_values,
                start=1,
            ):
                field = _field_for_header(cell)
                if field and field not in mapping:
                    mapping[field] = column_number
            if REQUIRED_FIELDS.issubset(mapping):
                return sheet, row_number, mapping
    return None, 0, {}


def parse_gantt_workbook(file_or_bytes) -> GanttImportResult:
    if hasattr(file_or_bytes, "read"):
        content = file_or_bytes.read()
        if hasattr(file_or_bytes, "seek"):
            file_or_bytes.seek(0)
    else:
        content = bytes(file_or_bytes)
    if len(content) > MAX_WORKBOOK_BYTES:
        return GanttImportResult(
            sheet_name="",
            header_row=0,
            rows=[],
            warnings=[],
            errors=["The Gantt workbook must be 10 MB or smaller."],
        )
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ARCHIVE_MEMBERS:
                raise ValueError("The workbook contains too many internal files.")
            if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                raise ValueError("The workbook expands beyond the 100 MB safety limit.")
        workbook = load_workbook(
            io.BytesIO(content),
            data_only=True,
            read_only=True,
            keep_links=False,
        )
    except (zipfile.BadZipFile, InvalidFileException, OSError, ValueError) as error:
        return GanttImportResult(
            sheet_name="",
            header_row=0,
            rows=[],
            warnings=[],
            errors=[f"The Excel workbook could not be read safely: {error}"],
        )
    if len(workbook.worksheets) > MAX_WORKSHEETS:
        return GanttImportResult(
            sheet_name="",
            header_row=0,
            rows=[],
            warnings=[],
            errors=[f"The Gantt workbook may contain at most {MAX_WORKSHEETS} worksheets."],
        )
    sheet, header_row, columns = _find_header(workbook)
    if sheet is None:
        return GanttImportResult(
            sheet_name="",
            header_row=0,
            rows=[],
            warnings=[],
            errors=[
                "No Gantt table was found. Include Task, Start Date, and End Date columns."
            ],
        )

    rows = []
    warnings = []
    errors = []
    current_phase = ""
    empty_run = 0
    data_rows = sheet.iter_rows(
        min_row=header_row + 1,
        max_row=1000,
        max_col=max(columns.values()),
        values_only=False,
    )
    for row_number, row_cells in enumerate(data_rows, start=header_row + 1):
        row_dimensions = getattr(sheet, "row_dimensions", None)
        if row_dimensions and row_dimensions[row_number].hidden:
            continue
        values = {
            field: row_cells[column_number - 1].value
            for field, column_number in columns.items()
        }
        task = str(values.get("task") or "").strip()
        if not task:
            empty_run += 1
            if empty_run >= MAX_EMPTY_TASK_ROWS:
                break
            continue
        empty_run = 0
        if task.casefold().startswith(IGNORED_TASK_PREFIXES):
            continue

        start_date = _parse_date(values.get("start"), workbook)
        end_date = _parse_date(values.get("end"), workbook)
        assigned_to = str(values.get("assigned_to") or "").strip()
        progress = _parse_progress(values.get("progress"))
        explicit_phase = str(values.get("phase") or "").strip()

        if not start_date and not end_date and not assigned_to and progress is None:
            current_phase = explicit_phase or task
            continue
        if not start_date or not end_date:
            errors.append(
                f"Row {row_number}: '{task}' needs both a start date and an end date."
            )
            continue
        if end_date < start_date:
            errors.append(
                f"Row {row_number}: '{task}' ends before its start date."
            )
            continue
        if progress is None and values.get("progress") not in (None, ""):
            warnings.append(
                f"Row {row_number}: progress for '{task}' was not recognized and was set to 0%."
            )

        progress = progress or 0.0
        rows.append(
            {
                "source_row": row_number,
                "milestone_id": str(values.get("milestone_id") or "").strip(),
                "phase": explicit_phase or current_phase,
                "task": task,
                "assigned_to": assigned_to,
                "progress_percent": round(progress, 1),
                "start_date": start_date.isoformat(),
                "due_date": end_date.isoformat(),
                "status": _status(values.get("status"), progress),
                "next_action": str(values.get("next_action") or "").strip(),
            }
        )

    if not rows and not errors:
        errors.append("The Gantt table does not contain any importable task rows.")
    return GanttImportResult(
        sheet_name=sheet.title,
        header_row=header_row,
        rows=rows,
        warnings=warnings,
        errors=errors,
    )


def resolve_gantt_rows(
    parsed_rows,
    *,
    project,
    members,
    default_owner_member_id,
    updated_at,
):
    member_lookup = {}
    for member in members:
        member_id = str(member.get("member_id", "")).strip()
        for value in (
            member_id,
            member.get("email", ""),
            member.get("name", ""),
            member.get("display_name", ""),
        ):
            key = str(value or "").strip().casefold()
            if key:
                member_lookup[key] = member_id

    warnings = []
    output = []
    duplicate_counts = {}
    for row in parsed_rows:
        assigned_to = str(row.get("assigned_to") or "").strip()
        owner_member_id = member_lookup.get(
            assigned_to.casefold(),
            default_owner_member_id,
        )
        if assigned_to and assigned_to.casefold() not in member_lookup:
            warnings.append(
                f"Row {row['source_row']}: '{assigned_to}' was not found in the lab roster; "
                "the selected default owner will be used."
            )
        signature = "|".join(
            [
                project["project_id"],
                row["phase"],
                row["task"],
                row["start_date"],
                row["due_date"],
            ]
        )
        duplicate_counts[signature] = duplicate_counts.get(signature, 0) + 1
        digest = hashlib.sha256(
            f"{signature}|{duplicate_counts[signature]}".encode()
        ).hexdigest()[:12].upper()
        output.append(
            {
                "milestone_id": f"MS-GANTT-{digest}",
                "project_id": project["project_id"],
                "project": project.get("project", ""),
                "aim": project.get("aim", ""),
                "milestone": row["task"],
                "time_window": row["phase"],
                "owner_member_id": owner_member_id,
                "start_date": row["start_date"],
                "status": row["status"],
                "review_status": "Pending",
                "next_action": row["next_action"] or "Review imported Gantt task",
                "due_date": row["due_date"],
                "blocker_reason": "",
                "help_needed_from": "",
                "progress_percent": str(row["progress_percent"]),
                "updated_at": updated_at,
            }
        )
    return output, warnings


def merge_project_gantt(existing_rows, project_id, imported_rows):
    preserved = [
        row
        for row in existing_rows
        if not (
            row.get("project_id") == project_id
            and str(row.get("milestone_id", "")).startswith("MS-GANTT-")
        )
    ]
    return [*preserved, *imported_rows]


def build_gantt_context(project, milestones, member_lookup):
    if not project:
        return {"project": None, "rows": [], "ticks": []}
    rows = []
    for milestone in milestones:
        if milestone.get("project_id") != project.get("project_id"):
            continue
        try:
            start_date = date.fromisoformat(str(milestone.get("start_date", "")))
            due_date = date.fromisoformat(str(milestone.get("due_date", "")))
        except ValueError:
            continue
        if due_date < start_date:
            continue
        progress = _parse_progress(milestone.get("progress_percent"))
        if progress is None:
            progress = {
                "Completed": 100.0,
                "In progress": 50.0,
                "Blocked": 25.0,
            }.get(milestone.get("status"), 0.0)
        rows.append(
            {
                **milestone,
                "start": start_date,
                "end": due_date,
                "progress": progress,
                "owner": member_lookup.get(
                    milestone.get("owner_member_id"),
                    milestone.get("owner_member_id", ""),
                ),
                "phase": milestone.get("time_window") or "Unassigned phase",
            }
        )
    if not rows:
        return {"project": project, "rows": [], "ticks": []}

    chart_start = min(row["start"] for row in rows)
    chart_end = max(row["end"] for row in rows)
    total_days = max((chart_end - chart_start).days + 1, 1)
    for row in rows:
        row["left"] = round((row["start"] - chart_start).days / total_days * 100, 4)
        row["width"] = round(
            max((row["end"] - row["start"]).days + 1, 1) / total_days * 100,
            4,
        )
        row["progress_width"] = round(row["progress"], 1)

    tick_count = min(6, total_days)
    ticks = []
    for index in range(tick_count):
        offset = round(index * (total_days - 1) / max(tick_count - 1, 1))
        ticks.append(
            {
                "label": (
                    date.fromordinal(chart_start.toordinal() + offset)
                ).strftime("%b %d").replace(" 0", " "),
                "left": round(offset / total_days * 100, 4),
            }
        )
    return {
        "project": project,
        "rows": sorted(rows, key=lambda row: (row["start"], row["end"], row["milestone"])),
        "ticks": ticks,
        "start_date": chart_start,
        "end_date": chart_end,
        "total_days": total_days,
    }
