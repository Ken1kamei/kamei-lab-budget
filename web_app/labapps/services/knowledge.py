import re
import zipfile
import csv
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree

import pdfplumber
from openpyxl import load_workbook


WORD_NAMESPACE = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
WORD = {"w": WORD_NAMESPACE}
WORD_VALUE = f"{{{WORD_NAMESPACE}}}val"
MAX_DOCUMENT_XML_BYTES = 15 * 1024 * 1024
MAX_KNOWLEDGE_FILE_BYTES = 25 * 1024 * 1024
MAX_PDF_PAGES = 200
MAX_EXTRACTED_CHARACTERS = 2_000_000
MAX_BLOCKS = 5_000
MAX_TABLE_ROWS = 1_000
EXTRACTED_METADATA_KEYS = {
    "document_title",
    "summary",
    "sections",
    "overview",
    "materials",
    "procedure",
    "parse_status",
    "parser",
    "section_count",
    "parse_message",
    "last_reprocess_status",
    "last_reprocess_error",
    "last_reprocess_at",
}
KNOWN_HEADINGS = {
    "protocol metadata",
    "purpose",
    "purpose / overview",
    "overview",
    "safety",
    "safety / ethics notes",
    "materials",
    "preparation before starting",
    "procedure",
    "quality control",
    "quality control / acceptance criteria",
    "data table",
    "data storage and file naming",
    "troubleshooting",
    "references",
    "version history",
}


def extract_knowledge_metadata(filename, content, content_type=""):
    if len(content) > MAX_KNOWLEDGE_FILE_BYTES:
        return {
            "parse_status": "failed",
            "parser": "size-check",
            "parse_message": "The document exceeds the 25 MB extraction limit.",
            "content_type": content_type,
        }
    suffix = Path(filename).suffix.casefold()
    parser = {
        ".docx": "docx-v1",
        ".pdf": "pdf-v1",
        ".md": "text-v1",
        ".txt": "text-v1",
        ".csv": "csv-v1",
        ".xlsx": "xlsx-v1",
        ".pptx": "pptx-v1",
    }.get(suffix, "unsupported")
    try:
        if suffix == ".docx":
            parsed = _parse_docx(content)
        elif suffix == ".pdf":
            parsed = _parse_pdf(content)
        elif suffix in {".md", ".txt"}:
            parsed = _parse_text(content)
        elif suffix == ".csv":
            parsed = _parse_csv(content)
        elif suffix == ".xlsx":
            parsed = _parse_xlsx(content)
        elif suffix == ".pptx":
            parsed = _parse_pptx(content)
        else:
            return {
                "parse_status": "unsupported",
                "parser": parser,
                "parse_message": "Automatic text extraction is available for DOCX, PDF, MD, TXT, CSV, XLSX, and PPTX files.",
            }
        parsed.update(
            {
                "parse_status": "parsed",
                "parser": parser,
                "section_count": len(parsed.get("sections", [])),
            }
        )
        return parsed
    except Exception as error:
        return {
            "parse_status": "failed",
            "parser": parser,
            "parse_message": _clean_text(str(error))[:300]
            or "The document could not be parsed.",
            "content_type": content_type,
        }


def _parse_docx(content):
    if not zipfile.is_zipfile(BytesIO(content)):
        raise ValueError("The uploaded DOCX file is not a valid Word document.")
    with zipfile.ZipFile(BytesIO(content)) as archive:
        try:
            document_info = archive.getinfo("word/document.xml")
        except KeyError as error:
            raise ValueError("The Word document does not contain document.xml.") from error
        if document_info.file_size > MAX_DOCUMENT_XML_BYTES:
            raise ValueError("The Word document is too large to extract safely.")
        if (
            document_info.compress_size
            and document_info.file_size / document_info.compress_size > 100
        ):
            raise ValueError("The Word document has an unsafe compression ratio.")
        root = ElementTree.fromstring(archive.read(document_info))

    body = root.find("w:body", WORD)
    if body is None:
        raise ValueError("The Word document body is empty.")

    blocks = []
    extracted_characters = 0
    for child in body:
        tag = child.tag.rsplit("}", 1)[-1]
        if tag == "p":
            text = _word_text(child)
            if not text:
                continue
            extracted_characters += len(text)
            if extracted_characters > MAX_EXTRACTED_CHARACTERS:
                raise ValueError("The Word document contains too much text to extract safely.")
            style_node = child.find("./w:pPr/w:pStyle", WORD)
            style = style_node.get(WORD_VALUE, "") if style_node is not None else ""
            if style.casefold().startswith("heading") or text.casefold() in KNOWN_HEADINGS:
                kind = "heading"
            elif style.casefold().startswith("listnumber"):
                kind = "numbered"
            elif style.casefold().startswith("listbullet"):
                kind = "bullets"
            else:
                kind = "paragraph"
            blocks.append({"kind": kind, "text": text})
        elif tag == "tbl":
            rows = []
            for row in child.findall("./w:tr", WORD):
                if len(rows) >= MAX_TABLE_ROWS:
                    raise ValueError("A Word table exceeds the safe row limit.")
                cells = []
                for cell in row.findall("./w:tc", WORD):
                    cell_lines = [
                        _word_text(paragraph)
                        for paragraph in cell.findall(".//w:p", WORD)
                    ]
                    cells.append("\n".join(line for line in cell_lines if line))
                if any(cells):
                    rows.append(cells)
            if rows:
                blocks.append({"kind": "table", "rows": rows})
        if len(blocks) > MAX_BLOCKS:
            raise ValueError("The Word document contains too many content blocks.")
    return _build_document(blocks)


def _parse_pdf(content):
    blocks = []
    extracted_characters = 0
    with pdfplumber.open(BytesIO(content)) as document:
        page_count = len(document.pages)
        if not page_count:
            raise ValueError("The PDF does not contain any pages.")
        for page in document.pages[:MAX_PDF_PAGES]:
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            for line in text.splitlines():
                line = _clean_text(line)
                if not line:
                    continue
                extracted_characters += len(line)
                if extracted_characters > MAX_EXTRACTED_CHARACTERS:
                    raise ValueError("The PDF contains too much text to extract safely.")
                kind = "heading" if _looks_like_heading(line) else "paragraph"
                blocks.append({"kind": kind, "text": line})
                if len(blocks) > MAX_BLOCKS:
                    raise ValueError("The PDF contains too many content blocks.")
    if not blocks:
        raise ValueError("No selectable text was found in the PDF.")
    parsed = _build_document(blocks)
    if page_count > MAX_PDF_PAGES:
        parsed["parse_message"] = f"Only the first {MAX_PDF_PAGES} PDF pages were extracted."
    return parsed


def _parse_text(content):
    text = content.decode("utf-8-sig")
    if len(text) > MAX_EXTRACTED_CHARACTERS:
        raise ValueError("The text document contains too much text to extract safely.")
    blocks = []
    for raw_line in text.splitlines():
        line = _clean_text(raw_line.lstrip("#").strip())
        if not line:
            continue
        if raw_line.lstrip().startswith("#") or _looks_like_heading(line):
            kind = "heading"
        elif re.match(r"^\d+[.)]\s+", line):
            kind = "numbered"
            line = re.sub(r"^\d+[.)]\s+", "", line)
        elif re.match(r"^[-*]\s+", line):
            kind = "bullets"
            line = re.sub(r"^[-*]\s+", "", line)
        else:
            kind = "paragraph"
        blocks.append({"kind": kind, "text": line})
        if len(blocks) > MAX_BLOCKS:
            raise ValueError("The text document contains too many content blocks.")
    if not blocks:
        raise ValueError("The text document is empty.")
    return _build_document(blocks)


def _parse_csv(content):
    text = content.decode("utf-8-sig")
    rows = [
        [_clean_text(cell) for cell in row]
        for row in csv.reader(text.splitlines())
        if any(_clean_text(cell) for cell in row)
    ]
    if not rows:
        raise ValueError("The CSV file is empty.")
    if len(rows) > MAX_TABLE_ROWS:
        raise ValueError("The CSV file exceeds the safe row limit.")
    return _build_document(
        [
            {"kind": "heading", "text": "CSV data"},
            {"kind": "table", "rows": rows},
        ]
    )


def _parse_xlsx(content):
    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
        data_only=True,
    )
    blocks = []
    extracted_characters = 0
    for worksheet in workbook.worksheets[:50]:
        blocks.append({"kind": "heading", "text": worksheet.title})
        rows = []
        for row_index, row in enumerate(
            worksheet.iter_rows(max_col=100, values_only=True),
            start=1,
        ):
            if row_index > MAX_TABLE_ROWS:
                raise ValueError("An Excel worksheet exceeds the safe row limit.")
            values = [_clean_text(value) for value in row]
            while values and not values[-1]:
                values.pop()
            if not any(values):
                continue
            extracted_characters += sum(len(value) for value in values)
            if extracted_characters > MAX_EXTRACTED_CHARACTERS:
                raise ValueError("The Excel workbook contains too much text to extract safely.")
            rows.append(values)
        if rows:
            blocks.append({"kind": "table", "rows": rows})
        if len(blocks) > MAX_BLOCKS:
            raise ValueError("The Excel workbook contains too many content blocks.")
    if not blocks:
        raise ValueError("The Excel workbook is empty.")
    return _build_document(blocks)


def _parse_pptx(content):
    if not zipfile.is_zipfile(BytesIO(content)):
        raise ValueError("The uploaded PPTX file is not a valid presentation.")
    blocks = []
    extracted_characters = 0
    with zipfile.ZipFile(BytesIO(content)) as archive:
        slide_names = sorted(
            (
                name
                for name in archive.namelist()
                if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
            ),
            key=lambda name: int(re.search(r"(\d+)", name.rsplit("/", 1)[-1]).group(1)),
        )
        for slide_number, name in enumerate(slide_names[:500], start=1):
            info = archive.getinfo(name)
            if info.file_size > MAX_DOCUMENT_XML_BYTES:
                raise ValueError("A PowerPoint slide is too large to extract safely.")
            root = ElementTree.fromstring(archive.read(info))
            texts = [
                _clean_text(node.text)
                for node in root.iter()
                if node.tag.rsplit("}", 1)[-1] == "t" and _clean_text(node.text)
            ]
            if not texts:
                continue
            extracted_characters += sum(len(text) for text in texts)
            if extracted_characters > MAX_EXTRACTED_CHARACTERS:
                raise ValueError("The presentation contains too much text to extract safely.")
            blocks.append(
                {
                    "kind": "heading",
                    "text": texts[0] or f"Slide {slide_number}",
                }
            )
            blocks.extend({"kind": "paragraph", "text": text} for text in texts[1:])
    if not blocks:
        raise ValueError("No readable text was found in the presentation.")
    return _build_document(blocks)


def _build_document(blocks):
    preamble = []
    sections = []
    current = None
    for block in blocks:
        if block["kind"] == "heading":
            current = {"heading": block["text"], "blocks": []}
            sections.append(current)
            continue
        if current is None:
            if block["kind"] == "paragraph":
                preamble.append(block["text"])
            elif block["kind"] == "table":
                current = {"heading": "Document details", "blocks": []}
                sections.append(current)
            else:
                current = {"heading": "Document summary", "blocks": []}
                sections.append(current)
        if current is not None:
            _append_block(current["blocks"], block)

    if not preamble and not sections:
        raise ValueError("No readable content was found in the document.")

    parsed = {
        "document_title": preamble[0] if preamble else "",
        "summary": preamble[1:] if len(preamble) > 1 else [],
        "sections": sections,
    }
    for key, aliases in {
        "overview": ("purpose", "overview"),
        "materials": ("materials",),
        "procedure": ("procedure",),
    }.items():
        parsed[key] = _flatten_sections(sections, aliases)
    return parsed


def _append_block(target, block):
    kind = block["kind"]
    if kind in {"bullets", "numbered"}:
        if target and target[-1].get("kind") == kind:
            target[-1]["items"].append(block["text"])
        else:
            target.append({"kind": kind, "items": [block["text"]]})
        return
    if kind == "table":
        target.append({"kind": "table", "rows": block["rows"]})
        return
    target.append({"kind": "paragraph", "text": block["text"]})


def _flatten_sections(sections, aliases):
    values = []
    for section in sections:
        heading = section["heading"].casefold()
        if not any(alias in heading for alias in aliases):
            continue
        for block in section["blocks"]:
            if block["kind"] == "paragraph":
                values.append(block["text"])
            elif block["kind"] in {"bullets", "numbered"}:
                values.extend(block["items"])
            elif block["kind"] == "table":
                values.extend(
                    " | ".join(cell for cell in row if cell)
                    for row in block["rows"]
                    if any(row)
                )
    return values


def _looks_like_heading(line):
    normalized = line.casefold().rstrip(":")
    if normalized in KNOWN_HEADINGS:
        return True
    return (
        len(line) <= 80
        and line.endswith(":")
        and len(line.split()) <= 10
    )


def _word_text(element):
    return _clean_text(
        "".join(
            node.text or ""
            for node in element.findall(".//w:t", WORD)
        )
    )


def _clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()
